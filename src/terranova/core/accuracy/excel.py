"""Write an accuracy report to an .xlsx workbook.

Single sheet for now (multi-sheet would be useful later for multi-run
comparisons).  Layout matches the standard remote-sensing format from
Congalton & Green 2019:

- Header block: scene name, run date, OA, kappa, n_samples.
- Confusion matrix with row totals (`Σ row`) and column totals (`Σ col`).
- Per-class block: predicted-class | user's accuracy | producer's
  accuracy | F1.

Uses openpyxl (ships with QGIS via pandas/processing) and falls back
to a clear error if it isn't importable.
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import TYPE_CHECKING

if TYPE_CHECKING:  # pragma: no cover
    from .metrics import AccuracyReport


def write_excel_report(
    report: "AccuracyReport",
    out_path: Path,
    *,
    raster_name: str = "",
    vector_name: str = "",
) -> Path:
    """Render ``report`` into an .xlsx at ``out_path``.  Returns the path."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError(
            "openpyxl isn't installed in this QGIS Python — install it via "
            "OSGeo4W Shell: `pip install openpyxl`, then retry."
        ) from exc

    wb = Workbook()
    ws = wb.active
    ws.title = "Accuracy"

    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="DDDDDD")
    diag_fill = PatternFill("solid", fgColor="E8F4E5")
    center = Alignment(horizontal="center", vertical="center")

    row = 1
    ws.cell(row=row, column=1, value="Terranova accuracy report").font = Font(
        bold=True, size=14
    )
    row += 1
    ws.cell(row=row, column=1, value="Generated").font = bold
    ws.cell(row=row, column=2, value=datetime.utcnow().isoformat(timespec="seconds"))
    row += 1
    if raster_name:
        ws.cell(row=row, column=1, value="Classified raster").font = bold
        ws.cell(row=row, column=2, value=raster_name)
        row += 1
    if vector_name:
        ws.cell(row=row, column=1, value="Validation vector").font = bold
        ws.cell(row=row, column=2, value=vector_name)
        row += 1
    ws.cell(row=row, column=1, value="Samples").font = bold
    ws.cell(row=row, column=2, value=int(report.n_samples))
    row += 1
    ws.cell(row=row, column=1, value="Overall accuracy").font = bold
    ws.cell(
        row=row,
        column=2,
        value=f"{float(report.overall_accuracy) * 100:.1f}%",
    )
    row += 1
    ws.cell(row=row, column=1, value="Kappa (Cohen)").font = bold
    ws.cell(row=row, column=2, value=round(float(report.kappa), 3))
    row += 2

    # Confusion matrix.
    classes = report.class_labels
    cm = report.confusion_matrix

    ws.cell(row=row, column=1, value="Confusion matrix").font = Font(bold=True, size=12)
    row += 1
    cm_header_row = row
    ws.cell(row=row, column=1, value="predicted ↓ / truth →").font = bold
    for j, cls in enumerate(classes):
        c = ws.cell(row=row, column=2 + j, value=int(cls))
        c.font = bold
        c.fill = header_fill
        c.alignment = center
    ws.cell(row=row, column=2 + len(classes), value="Σ row").font = bold
    row += 1
    row_totals = cm.sum(axis=1)
    col_totals = cm.sum(axis=0)
    for i, cls in enumerate(classes):
        c = ws.cell(row=row, column=1, value=int(cls))
        c.font = bold
        c.fill = header_fill
        c.alignment = center
        for j in range(len(classes)):
            cell = ws.cell(row=row, column=2 + j, value=int(cm[i, j]))
            cell.alignment = center
            if i == j:
                cell.fill = diag_fill
        ws.cell(row=row, column=2 + len(classes), value=int(row_totals[i])).font = bold
        row += 1
    # Column totals row.
    ws.cell(row=row, column=1, value="Σ col").font = bold
    for j in range(len(classes)):
        ws.cell(row=row, column=2 + j, value=int(col_totals[j])).font = bold
    ws.cell(row=row, column=2 + len(classes), value=int(cm.sum())).font = bold
    row += 2

    # Per-class metrics.
    ws.cell(row=row, column=1, value="Per-class metrics").font = Font(bold=True, size=12)
    row += 1
    for col_i, label in enumerate(
        ["Class", "User's accuracy", "Producer's accuracy", "F1"], start=1
    ):
        c = ws.cell(row=row, column=col_i, value=label)
        c.font = bold
        c.fill = header_fill
        c.alignment = center
    row += 1
    for i, cls in enumerate(classes):
        ws.cell(row=row, column=1, value=int(cls)).font = bold
        ws.cell(row=row, column=2, value=_pct(report.users_accuracy[i]))
        ws.cell(row=row, column=3, value=_pct(report.producers_accuracy[i]))
        ws.cell(row=row, column=4, value=_round(report.f1_per_class[i]))
        row += 1

    # Column widths — first column wide enough for the longest header
    # ("predicted ↓ / truth →"), the rest just snug to numbers.
    ws.column_dimensions["A"].width = 26
    for j in range(2, 2 + len(classes) + 1):
        ws.column_dimensions[get_column_letter(j)].width = 12

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    return out_path


def _round(v: float, ndigits: int = 4) -> float | str:
    """Round; render NaN as 'n/a' instead of pollute the cell."""
    import math

    if v is None or (isinstance(v, float) and math.isnan(v)):
        return "n/a"
    return round(float(v), ndigits)


def _pct(v: float) -> str:
    """Render an accuracy in [0,1] as a one-decimal percentage."""
    import math

    if v is None or (isinstance(v, float) and math.isnan(v)):
        return "n/a"
    return f"{float(v) * 100:.1f}%"
